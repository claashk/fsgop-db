# -*- coding: utf-8 -*-

import csv
from collections import namedtuple
import re
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    WITH_XLSX_SUPPORT = True
except ImportError:
    WITH_XLSX_SUPPORT = False


class Xlsx2Csv(object):
    """Adapter for Excel files to expose same interface as csv reader / writer

    Arguments:
        path (str): path to Excel xlsv file
        sheet (str): Name of sheet to open. If not provided, the first data sheet
            will be read
        read_only (bool): Passed verbatim to :func:`openpyxl.load_workbook`
        **kwargs (dict): Keyword arguments passed verbatim to
             :func:`openpyxl.load_workbook`
    """

    NUMBER_FORMAT_PATTERN = re.compile(r"([#0?]*)(?:\.([#0?]*)(E\+0+)?)?")

    def __init__(self, path=None, sheet=None, read_only=None, **kwargs):
        self._wb = None
        self._ws = None
        self._path = None
        self._read_only = True

        if path is not None:
            self.open(path, sheet=sheet, read_only=read_only, **kwargs)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def __iter__(self):
        """Iterate over all rows of the input file

        Yield:
            tuple: Tuple of strings containing the content of each cell
        """
        yield from self.iter_rows()

    def open(self, path, sheet=None, read_only=None, **kwargs):
        if not WITH_XLSX_SUPPORT:
            raise RuntimeError("Missing package: openpyxl")
        self._path = Path(path)

        if not self._path.exists():
            self._read_only = bool(read_only) if read_only is not None else False
            self._wb = Workbook()
        else:
            self._read_only = bool(read_only) if read_only is not None else True
            self._wb = load_workbook(path, read_only=self._read_only, **kwargs)
        self.open_sheet(sheet)

    def close(self):
        self._ws = None
        if self._wb is not None:
            if not self._read_only:
                self._wb.save(str(self._path))
            self._wb.close()
            self._wb = None
            self._path = None

    def open_sheet(self, sheet=None):
        if self._wb is None:
            raise RuntimeError("No workbook open")
        if sheet is not None:
            self._ws = self._wb.get_sheet_by_name(sheet)
        else:
            self._ws = self._wb.worksheets[0]

    def writerow(self, row):
        self._ws.append(row)

    def iter_rows(self):
        if self._wb is None or self._ws is None:
            return
        for row in self._ws.rows:
            yield tuple(self.cell_to_str(cell) for cell in row)

    @classmethod
    def cell_to_str(cls, cell):
        if cell.value is None:
            return ""
        return cls.cell_format(cell).format(cell.value)

    @classmethod
    def cell_format(cls, cell):
        """Convert Excel number format to python format string

        Arguments:
            cell (openpyxl.Cell): Cell

        Return:
            str: Format string containing the python format for ``cell``
        """
        if cell.data_type != 'n' or cell.number_format is None:
            return "{}"

        match = cls.NUMBER_FORMAT_PATTERN.match(cell.number_format)
        if match is None or match.start() == match.end():
            return "{}"

        lead, decimal, exp = match.groups()

        fmt = ['{:']
        if lead:
            if lead.startswith("0"):
                fmt.append("0")
            fmt.append(str(len(lead)))

        if decimal is not None:
            fmt.append(".{}".format(len(decimal)))
            if exp is not None:
                fmt.append("E")
            else:
                fmt.append('f')
        else:
            fmt.append("d")
        fmt.append("}")

        return "".join(fmt)


class CsvParser(object):
    """Simple CSV parser based on the csv library

    Creates a parser object, which can be used to parse CSV files of similar
    structure. The csv files parseable by this class need to be of the following
    structure:

    - A header containing zero or more fields in a key - value pattern

    - Either a fixed number of header rows or one row with column headings
      following a known pattern

    - An arbitrary number of data rows with identical number of columns. If a
      heading line is present, the number of heading columns must match the
      number of columns in each row.

    Rows containing column headings or header information are identified by
    regular expression patterns. Each of these patterns is matched against an
    entire row, obtained by concatenating all fields of a row with a blank space
    between fields. Note that a regex search is used, not a regex match. Thus only
    parts of a line have to match.

    Arguments:
        headings(list): Regular expression patterns used to identify column
            captions. Each pattern is matched against each row of the file, where
            separators are replaced by spaces.
        header (dict): Dictionary containing name and regular expression for
            header fields. Each regular expression is matched against each row
            of the file, where separators are replaced by spaces.
        translate(dict): Translations for headings. Each key matching a
            column heading will be replaced by the corresponding value in the
            parsed records

    Attributes:
        headings (list): Regular expression patterns used to identify a row
            containing column headings.
        header_fields (dict): Contains a unique key string and a matching
            :class:`re.Pattern` for each header field to extract.
        header (dict): After parsing this contains the extracted header
            information. For each key in :attr:`header_fields`, which is
            associated with a matching regular expression, the key is added to
            this dictionary with the value returned by `meth:`re.Match.groups`.
        row_type (namedtuple or list): If a heading row was identified, then this
            will be a namedtuple type containing the column headings. If no
            heading line is found, then this is the type :class:`list`.
    """
    def __init__(self, headings=None, header=None, translate=None, force_lowercase=False):
        self.translate = dict(translate) if translate is not None else dict()
        self.force_lowercase = bool(force_lowercase)
        if self.force_lowercase:
            self.translate = {k.lower(): v.lower()
                              for k, v in self.translate.items()}
        flags = re.IGNORECASE if force_lowercase else 0
        self.headings = None if headings is None else [re.compile(c, flags=flags)
                                                       for c in headings]
        if header is None:
            self.header_fields = dict()
        else:
            self.header_fields = {k: re.compile(v) for k, v in header.items()}

        self.header = dict()
        self.row_type = None
        self._cur = None
        self._line_number = -1

    def __call__(self, *args, **kwargs):
        """Invoke the parser

        Arguments:
            *args: Positional arguments passed verbatim to :meth:`parse`.
            **kwargs: Keyword arguments passed verbatim to :meth:`parse`.

        Yield:
            object: Same as :meth:`parse`
        """
        yield from self.parse(*args, **kwargs)

    @property
    def last_record(self):
        return self._cur

    @property
    def line_number(self):
        return self._line_number

    def parse_header(self, reader, max_rows=-1):
        """Extract header information from CSV file

        Attempts to extract header fields from the file until either the maximum
        number of header rows is reached or the row matches the heading pattern.
        The user can specify a regular expression for each header field to be
        parsed. Header fields not matching any regular expression are ignored.

        Arguments:
            reader (csv.reader): CSV reader object. The method will advance the
                reader, so that the next call of ``__next__`` will return the
                first row containing data.
            max_rows (int): Maximum number of header rows to extract. Parsing
                of the header will stop, if after `max_rows` have been read.
                If this value is negative, parsing will continue until a heading
                line is found.

        Raise:
            RuntimeError: If end of file is reached during parsing of the
            header.

        Return:
            tuple: dictionary with header information as stored in
            :attr:`CsvParser.header` and the row type stored in
            `CsvParser.row_type`
        """
        header = dict()
        row_type = self._tuple if self.row_type is None else self.row_type
        if self.headings is None and max_rows < 0:
            raise ValueError("Either column headings or number of header lines "
                             "must be specified")
        while max_rows != 0:
            try:
                row = next(reader)
                max_rows -= 1
            except StopIteration:
                raise RuntimeError("No column designation found")

            row_str = " ".join(row).strip()
            if not row_str:
                continue

            if self.headings and all([p.search(row_str) for p in self.headings]):
                names = [s.replace(' ', '_').replace('-', '_') .replace(".", "")
                         for s in row]
                if self.force_lowercase:
                    names = [s.lower() for s in names]
                return header, namedtuple('RowType',
                                          [self.translate.get(s, s)
                                           for s in names])
            for key, pattern in self.header_fields.items():
                match = pattern.search(row_str)
                if match:
                    header[key] = list(match.groups())
                    break
        return header, row_type

    def iter_body(self, reader):
        for self._line_number, row in enumerate(reader, start=1):
            self._cur = self.row_type(*row)
            yield self._cur

    def parse(self,
              path,
              skip_rows=-1,
              sheet=None,
              fmt=None,
              encoding="utf-8",
              **kwargs):
        """Parses a csv file defining the VSC definition

        Parses the content of a csv file and attempts to extract creation time
        and version from the header.

        Arguments:
            path (str): Path to CSV file
            skip_rows (int): Skip leading rows. Passed verbatim to
            **kwargs: Keyword arguments passed verbatim to csv.reader

        Yield:
            namedtuple: A named tuple for each row, containing the fields
            described by the header names. Spaces are replaced by underscores.
        """
        if fmt is None and str(path).endswith(".xlsx"):
            fmt = "xlsx"
        else:
            fmt = "csv"

        if fmt == "xlsx":
            with Xlsx2Csv(path, sheet=sheet, **kwargs) as xlsx_file:
                reader = iter(xlsx_file)
                self.header, self.row_type = self.parse_header(reader, skip_rows)
                yield from self.iter_body(reader)
        else:
            with open(path, newline='', encoding=encoding) as csv_file:
                reader = csv.reader(csv_file, **kwargs)
                self.header, self.row_type = self.parse_header(reader, skip_rows)
                yield from self.iter_body(reader)
        return

    @staticmethod
    def _tuple(*args):
        return args
